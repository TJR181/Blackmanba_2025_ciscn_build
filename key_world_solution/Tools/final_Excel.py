import os
import json
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from tqdm import tqdm


def normalize_cve(cve):
    match = re.match(r"CVE-(\d+)-(\d+)", cve.upper())
    if match:
        year, number = match.groups()
        return int(year), int(number)
    return (0, 0)


def collect_cve_data_from_json(data_dir="./data"):
    all_data = []
    if not os.path.exists(data_dir):
        print(f"❌ 目录 {data_dir} 不存在！")
        return all_data
    files = [f for f in os.listdir(data_dir) if f.endswith(".json")]
    for file in tqdm(files, desc="文件读取中", ncols=100):
        full_path = os.path.join(data_dir, file)
        with open(full_path, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
                if isinstance(data, list):
                    all_data.extend(data)
            except json.JSONDecodeError:
                print(f"⚠️ JSON解析失败：{file}")
    return all_data


def read_existing_cve_ids_from_excel(excel_file):
    if not os.path.exists(excel_file):
        return set()
    wb = load_workbook(excel_file)
    ws = wb[wb.sheetnames[0]]
    return {row[4] for row in ws.iter_rows(min_row=2, values_only=True) if row[4]}


def jsons_to_excel(data_dir="./data", result_dir="./result/"):
    os.makedirs(result_dir, exist_ok=True)
    excel_file = os.path.join(result_dir, "网络硬件设备安全知识库.xlsx")

    existing_ids = read_existing_cve_ids_from_excel(excel_file)
    all_data = collect_cve_data_from_json(data_dir)
    new_data = [item for item in all_data if item.get("CVE编号") not in existing_ids]

    if not new_data:
        print("✅ 没有新增数据，Excel 已是最新。")
        return

    new_data.sort(
        key=lambda x: (
            x.get("设备品牌", ""),
            x.get("设备类型", ""),
            x.get("产品型号", ""),
            normalize_cve(x.get("CVE编号", "")),
        )
    )
    new_data.sort(key=lambda x: normalize_cve(x.get("CVE编号", "")), reverse=True)

    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "网络硬件设备安全知识库"
        headers = [
            "序号",
            "设备品牌",
            "设备类型",
            "产品型号",
            "CVE编号",
            "漏洞描述",
            "攻击向量",
            "厂商补丁链接",
            "受影响版本",
            "公开日期",
        ]
        ws.append(headers)
        font = Font(bold=True)
        fill = PatternFill("solid", fgColor="FFFF00")
        border = Border(*[Side(style="thin")] * 4)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    start_index = ws.max_row + 1
    for idx, item in enumerate(
        tqdm(new_data, desc="写入Excel", ncols=100), start=start_index
    ):
        row = [
            idx - 1,
            item.get("设备品牌", ""),
            item.get("设备类型", ""),
            item.get("产品型号", ""),
            item.get("CVE编号", ""),
            item.get("漏洞描述", ""),
            item.get("攻击向量", ""),
            "；".join(item.get("厂商补丁链接", [])),
            "，".join(item.get("受影响版本", [])),
            item.get("公开日期", ""),
        ]
        ws.append(row)
        for col in range(1, len(row) + 1):
            cell = ws.cell(row=idx, column=col)
            cell.border = Border(*[Side(style="thin")] * 4)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=False
            )

    col_widths = [6, 10, 10, 20, 20, 60, 12, 40, 30, 16]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(excel_file)
    print(f"✅ 已写入 {len(new_data)} 条记录到 {excel_file}，总共 {ws.max_row - 1} 条")
    print("🎉 完成更新！")
